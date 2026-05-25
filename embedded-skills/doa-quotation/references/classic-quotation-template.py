# -*- coding: utf-8 -*-
"""
经典报价 Excel 生成脚本骨架

用途：
- 生成对客商务版工时报价单（Service Quotation / 服务报价单）
- 双工作表结构：「项目报价」 + 「开发工作量」
- 适配通用报价单风格
- 左上角自动嵌入公司 Logo（从 brand-assets/ 目录读取）

参考模板：
- 示例客户.示例项目.工时报价.xlsx

使用方式：
1. 替换 CONFIG 和 WORKLOAD_ITEMS 中的占位数据
2. 运行：python classic-quotation-template.py
"""

import os
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─────────────────────────────────────────────
# 样式常量（与参考模板一致）
# ─────────────────────────────────────────────

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Sheet 1 使用微软雅黑，Sheet 2 使用 DengXian（等线）
FONT_TITLE = Font(name="微软雅黑", size=28, bold=False)               # Service Quotation 标题
FONT_SECTION = Font(name="微软雅黑", size=9, bold=True)               # "项目服务价格" / "CR"
FONT_HEADER = Font(name="微软雅黑", size=9.5, bold=True)              # 表头
FONT_BODY = Font(name="微软雅黑", size=9)                             # 正文 9pt
FONT_BODY_10 = Font(name="微软雅黑", size=10)                         # 正文 10pt
FONT_COMPANY = Font(name="微软雅黑", size=9)                          # 公司信息
FONT_LABEL = Font(name="微软雅黑", size=10)                           # To/From 标签
FONT_LABEL_BOLD = Font(name="微软雅黑", size=10.5, bold=True)         # To/From 标题
FONT_TOTAL = Font(name="微软雅黑", size=10.5, bold=True)              # 汇总行

# Sheet 2: 开发工作量 字体（等线）
FONT_WL = Font(name="DengXian", size=11)                              # 开发工作量正文
FONT_WL_BOLD = Font(name="DengXian", size=11, bold=True)              # 开发工作量加粗
FONT_WL_TITLE = Font(name="DengXian", size=12, bold=True)             # 工作量评估表标题
FONT_WL_SM = Font(name="DengXian", size=10)                           # 开发工作量小号

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

FILL_HEADER = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")

MONEY_FORMAT = r'\¥#,##0.00;[Red]\¥\-#,##0.00'
MONEY_FORMAT_WL = r'"¥"#,##0.00_);[Red]\("¥"#,##0.00\)'
PERCENT_FORMAT = "0%"
DATE_FORMAT = "mm-dd-yy"

# ─────────────────────────────────────────────
# 项目配置（替换为实际项目信息）
# ─────────────────────────────────────────────

CONFIG = {
    "output_file": "示例客户.示例项目.工时报价.xlsx",
    "project_name": "示例项目",
    "customer_name": "示例客户",

    # 报价方公司信息
    "vendor_name": "您的公司名称",
    "vendor_name_en": "Your Company Name",
    "vendor_address_cn": "Add：您的公司地址",
    "vendor_address_en": "Your Company Address",
    "vendor_tel": "TEL：+86 21 12345678",
    "vendor_fax": "FAX：+86 21 12345679",

    # 报价标题
    "quote_title": "Service Quotation\n服务报价单",

    # 客户方联系人
    "customer_contact": "待填写",
    "customer_company": "待填写",
    "customer_address": "",
    "customer_email": "",
    "customer_phone": "",

    # 报价方联系人
    "sales_contact": "待填写",
    "sales_email": "sales@example.com",
    "sales_phone": "138xxxx0000",

    # 日期与账期
    "quote_date": date.today(),
    "due_days": 30,

    # 税率与单价（单价必须由用户提供，此处为占位示例）
    "tax_rate": 0.06,
    "unit_price": 0,  # TODO: 由用户提供，例如 2200、3000、3500

    # CR 变更开发单价（与主单价一致，或由用户单独指定）
    "cr_unit_price": 0,  # TODO: 由用户提供

    # Logo 文件名（自动从 brand-assets/ 读取）
    "logo_file": "company_logo.png",
}

# ─────────────────────────────────────────────
# 工作量明细（替换为实际任务拆分）
# ─────────────────────────────────────────────

WORKLOAD_ITEMS = [
    {
        "scenario": "需求沟通",
        "task_name": "1、需求沟通；\n2、方案设计",
        "pm_ba_sa_days": 1,
        "pg_pt_days": 1,
        "remark": "",
    },
    {
        "scenario": "接口设计",
        "task_name": "1、接口设计",
        "pm_ba_sa_days": 0,
        "pg_pt_days": 1,
        "remark": "",
    },
    {
        "scenario": "开发实施",
        "task_name": "功能开发与联调",
        "pm_ba_sa_days": 1,
        "pg_pt_days": 5,
        "remark": "",
    },
    {
        "scenario": "测试",
        "task_name": "SIT测试",
        "pm_ba_sa_days": 0,
        "pg_pt_days": 1,
        "remark": "",
    },
    {
        "scenario": "上线",
        "task_name": "项目部署",
        "pm_ba_sa_days": 0,
        "pg_pt_days": 1,
        "remark": "",
    },
]

NOTES = [
    "1. 报价说明：报价包含6%增值税；",
    "2. 服务说明：非现场服务（或现场服务）；",
    "3. 知识产权：第三方知识产权归第三方所有；为客户定制开发的部分知识产权归属双方共同所有；",
    "4. 账期：Net30天；",
    "5. 付款方式：项目上线：100%；",
    "6. 免费维保期：项目上线验收后一个月；",
    "7. 违约说明：违约上限为合同金额的100%。",
]


# ─────────────────────────────────────────────
# 工具函数
# ─────────────────────────────────────────────

def get_output_file_name():
    return CONFIG.get("output_file") or f"{CONFIG['customer_name']}.{CONFIG['project_name']}.工时报价.xlsx"


def style_range(ws, cell_range, fill=None, font=None, alignment=None, border=None):
    """批量设置单元格样式"""
    for row in ws[cell_range]:
        for cell in row:
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment
            if border is not None:
                cell.border = border


def set_col_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def set_row_heights(ws, heights: dict):
    for row, height in heights.items():
        ws.row_dimensions[row].height = height


def find_logo_path():
    """从 brand-assets/ 目录查找 Logo 文件"""
    candidates = [
        Path(__file__).parent / "brand-assets" / CONFIG["logo_file"],
        Path(__file__).parent.parent / "brand-assets" / CONFIG["logo_file"],
        Path.home() / ".claude" / "skills" / "doa-quotation" / "brand-assets" / CONFIG["logo_file"],
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    return None


# ─────────────────────────────────────────────
# Sheet 1: 项目报价
# ─────────────────────────────────────────────

def build_quote_sheet(wb):
    ws = wb.active
    ws.title = "项目报价"

    # 列宽（与参考模板精确匹配）
    set_col_widths(ws, {
        "A": 9.62, "B": 4.38, "C": 6.38, "D": 4.62, "E": 4.25,
        "F": 6.62, "G": 13.25, "H": 16.75, "I": 13.62, "J": 4.62,
        "K": 11.62, "L": 15.88, "M": 22.0, "N": 12.38, "O": 9.12,
    })

    # 行高
    set_row_heights(ws, {
        1: 17.45, 2: 17.45, 3: 17.45, 4: 17.45, 5: 17.45, 6: 17.45,
        7: 20.1, 8: 20.1, 9: 20.1, 10: 20.1, 11: 20.1, 12: 20.1,
        13: 20.1, 14: 20.1, 15: 20.1, 16: 20.1, 17: 20.1, 18: 20.1,
        19: 9.75, 20: 20.1, 21: 20.1, 22: 20.1, 23: 20.1,
        24: 27.75, 25: 16.5, 26: 124.9,
    })

    # ── Logo 区域（A1:D6 合并，Logo 缩小显示在 A2 附近）──
    ws.merge_cells("A1:D6")
    logo_path = find_logo_path()
    if logo_path:
        logo = XlImage(logo_path)
        logo.width = 150
        logo.height = 35
        ws.add_image(logo, "A2")

    # ── 标题区域 ──
    ws.merge_cells("E1:K6")
    ws["E1"] = CONFIG["quote_title"]
    ws["E1"].font = FONT_TITLE
    ws["E1"].alignment = ALIGN_CENTER

    # ── 公司信息（右上角，右对齐）──
    company_rows = [
        ("L1", CONFIG["vendor_name"]),
        ("L2", CONFIG["vendor_name_en"]),
        ("L3", CONFIG["vendor_address_cn"]),
        ("L4", CONFIG["vendor_address_en"]),
        ("L5", CONFIG["vendor_tel"]),
        ("L6", CONFIG["vendor_fax"]),
    ]
    for merge_range in ("L1:N1", "L2:N2", "L3:N3", "L4:N4", "L5:N5", "L6:N6"):
        ws.merge_cells(merge_range)
    for cell_ref, val in company_rows:
        ws[cell_ref] = val
        ws[cell_ref].font = FONT_COMPANY
        ws[cell_ref].alignment = ALIGN_RIGHT

    # ── To / From 联系信息 ──
    ws["A7"] = "To："
    ws["A7"].font = FONT_LABEL_BOLD
    ws["I7"] = "From:"
    ws["I7"].font = FONT_LABEL_BOLD

    to_from_labels = [
        (8, "A", "Contact/联系人:", "I", "Contact/联系人:"),
        (9, "A", "Company/公司:", "I", "Email/邮箱:"),
        (10, "A", "Address/地址:", "I", "Phone/电话:"),
        (11, "A", "Email/邮箱:", "I", "Quo date/报价日期:"),
        (12, "A", "Phone/电话:", "I", "Due Date/截止日期:"),
    ]
    for row, col_l, label_l, col_r, label_r in to_from_labels:
        ws[f"{col_l}{row}"] = label_l
        ws[f"{col_l}{row}"].font = FONT_LABEL
        ws[f"{col_l}{row}"].alignment = ALIGN_LEFT
        ws[f"{col_r}{row}"] = label_r
        ws[f"{col_r}{row}"].font = FONT_LABEL
        ws[f"{col_r}{row}"].alignment = ALIGN_LEFT

    # To 区合并单元格 + 值
    for merge in ("A8:B8", "C8:G8", "A9:B9", "C9:G9", "A10:B10", "C10:G10",
                   "A11:B11", "C11:G11", "A12:B12", "C12:G12"):
        ws.merge_cells(merge)
    ws["C8"] = CONFIG["customer_contact"]
    ws["C9"] = CONFIG["customer_company"]
    ws["C10"] = CONFIG["customer_address"]
    ws["C11"] = CONFIG["customer_email"]
    ws["C12"] = CONFIG["customer_phone"]

    # From 区合并单元格 + 值
    for merge in ("I8:J8", "K8:N8", "I9:J9", "K9:N9", "I10:J10", "K10:N10",
                   "I11:J11", "K11:N11", "I12:J12", "K12:N12"):
        ws.merge_cells(merge)
    ws["K8"] = CONFIG["sales_contact"]
    ws["K9"] = CONFIG["sales_email"]
    ws["K10"] = CONFIG["sales_phone"]
    ws["K11"] = CONFIG["quote_date"]
    ws["K11"].number_format = DATE_FORMAT
    ws["K12"] = f'=IF(K11="","",K11+{CONFIG["due_days"]})'
    ws["K12"].number_format = DATE_FORMAT

    # To/From 区样式
    for r in range(7, 13):
        style_range(ws, f"A{r}:N{r}", font=FONT_LABEL, alignment=ALIGN_LEFT, border=THIN_BORDER)
    ws["A7"].font = FONT_LABEL_BOLD
    ws["I7"].font = FONT_LABEL_BOLD

    # ── 项目服务价格 Section ──
    ws.merge_cells("A13:N13")
    ws["A13"] = "项目服务价格"
    ws["A13"].font = FONT_SECTION
    ws["A13"].alignment = ALIGN_LEFT
    style_range(ws, "A13:N13", border=THIN_BORDER)

    # 表头行 14
    header_labels = {
        "A": "Service/服务", "D": "Description/描述",
        "H": "QTY/数量", "I": "Unit/单位", "J": "Unit Price/单价",
        "L": "Total/小计", "M": "Total Inc VAT/含税小计", "N": "Tax Rate税率",
    }
    ws.merge_cells("A14:C14")
    ws.merge_cells("D14:G14")
    ws.merge_cells("J14:K14")
    for col, val in header_labels.items():
        ws[f"{col}14"] = val
    style_range(ws, "A14:N14", fill=FILL_HEADER, font=FONT_HEADER, alignment=ALIGN_CENTER, border=THIN_BORDER)

    # 工作量表小计行位置
    subtotal_row_wl = 4 + len(WORKLOAD_ITEMS)
    unit_price = CONFIG["unit_price"]

    # PM/BA 行
    ws.merge_cells("A15:C15")
    ws["A15"] = "开发定制"
    ws.merge_cells("D15:G15")
    ws["D15"] = "项目管理/业务顾问(PM/BA)"
    ws["H15"] = f"=开发工作量!E{subtotal_row_wl}"
    ws["I15"] = '=IF(H15="","","Manday")'
    ws.merge_cells("J15:K15")
    ws["J15"] = unit_price
    ws["L15"] = "=H15*J15"
    ws["M15"] = '=IF(H15="","",ROUND(H15*J15+H15*J15*N15,2))'
    ws["N15"] = CONFIG["tax_rate"]
    style_range(ws, "A15:N15", font=FONT_BODY, alignment=ALIGN_CENTER, border=THIN_BORDER)
    ws["J15"].number_format = MONEY_FORMAT
    ws["J15"].alignment = ALIGN_RIGHT
    ws["L15"].number_format = MONEY_FORMAT
    ws["M15"].number_format = MONEY_FORMAT
    ws["N15"].number_format = PERCENT_FORMAT

    # PG/PT 行
    ws.merge_cells("A16:C16")
    ws["A16"] = "开发定制"
    ws.merge_cells("D16:G16")
    ws["D16"] = "程序开发/系统测试(PG/PT)"
    ws["H16"] = f"=开发工作量!F{subtotal_row_wl}"
    ws["I16"] = '=IF(H16="","","Manday")'
    ws.merge_cells("J16:K16")
    ws["J16"] = unit_price
    ws["L16"] = "=H16*J16"
    ws["M16"] = '=IF(H16="","",ROUND(H16*J16+H16*J16*N16,2))'
    ws["N16"] = CONFIG["tax_rate"]
    style_range(ws, "A16:N16", font=FONT_BODY, alignment=ALIGN_CENTER, border=THIN_BORDER)
    ws["J16"].number_format = MONEY_FORMAT
    ws["J16"].alignment = ALIGN_RIGHT
    ws["L16"].number_format = MONEY_FORMAT
    ws["M16"].number_format = MONEY_FORMAT
    ws["N16"].number_format = PERCENT_FORMAT

    # 未税小计
    ws.merge_cells("A17:K17")
    ws["A17"] = "未税小计(RMB)"
    ws["A17"].font = FONT_TOTAL
    ws["A17"].alignment = ALIGN_RIGHT
    ws.merge_cells("L17:N17")
    ws["L17"] = "=L15+L16"
    ws["L17"].font = FONT_TOTAL
    ws["L17"].number_format = MONEY_FORMAT
    ws["L17"].alignment = ALIGN_CENTER
    style_range(ws, "A17:N17", border=THIN_BORDER)

    # 含税合计
    ws.merge_cells("A18:K18")
    ws["A18"] = "Total含税合计(RMB)"
    ws["A18"].font = FONT_TOTAL
    ws["A18"].alignment = ALIGN_RIGHT
    ws.merge_cells("L18:N18")
    ws["L18"] = "=L17*1.06"
    ws["L18"].font = FONT_TOTAL
    ws["L18"].number_format = MONEY_FORMAT
    ws["L18"].alignment = ALIGN_CENTER
    style_range(ws, "A18:N18", border=THIN_BORDER)

    # 空行 19（间距）
    ws.merge_cells("A19:N19")

    # ── CR/Enhancement 费用 Section ──
    ws.merge_cells("A20:N20")
    ws["A20"] = "CR/Enhancement 费用"
    ws["A20"].font = FONT_SECTION
    ws["A20"].alignment = ALIGN_LEFT
    style_range(ws, "A20:N20", border=THIN_BORDER)

    # CR 表头行 21
    ws.merge_cells("A21:C21")
    ws.merge_cells("D21:G21")
    ws.merge_cells("J21:K21")
    for col, val in header_labels.items():
        ws[f"{col}21"] = val
    style_range(ws, "A21:N21", fill=FILL_HEADER, font=FONT_HEADER, alignment=ALIGN_CENTER, border=THIN_BORDER)

    # CR PM/BA 行
    cr_price = CONFIG["cr_unit_price"]
    ws.merge_cells("A22:C22")
    ws["A22"] = "开发定制"
    ws.merge_cells("D22:G22")
    ws["D22"] = "项目管理/业务顾问(PM/BA)"
    ws["I22"] = "Manday"
    ws.merge_cells("J22:K22")
    ws["J22"] = cr_price
    ws["L22"] = "=H22*J22"
    ws["M22"] = '=IF(H22="","",ROUND(H22*J22+H22*J22*N22,2))'
    ws["N22"] = CONFIG["tax_rate"]
    style_range(ws, "A22:N22", font=FONT_BODY, alignment=ALIGN_CENTER, border=THIN_BORDER)
    ws["J22"].number_format = MONEY_FORMAT
    ws["J22"].alignment = ALIGN_RIGHT
    ws["L22"].number_format = MONEY_FORMAT
    ws["M22"].number_format = MONEY_FORMAT
    ws["N22"].number_format = PERCENT_FORMAT

    # CR PG/PT 行
    ws.merge_cells("A23:C23")
    ws["A23"] = "开发定制"
    ws.merge_cells("D23:G23")
    ws["D23"] = "程序开发/系统测试(PG/PT)"
    ws["I23"] = "Manday"
    ws.merge_cells("J23:K23")
    ws["J23"] = cr_price
    ws["L23"] = "=H23*J23"
    ws["M23"] = '=IF(H23="","",ROUND(H23*J23+H23*J23*N23,2))'
    ws["N23"] = CONFIG["tax_rate"]
    style_range(ws, "A23:N23", font=FONT_BODY, alignment=ALIGN_CENTER, border=THIN_BORDER)
    ws["J23"].number_format = MONEY_FORMAT
    ws["J23"].alignment = ALIGN_RIGHT
    ws["L23"].number_format = MONEY_FORMAT
    ws["M23"].number_format = MONEY_FORMAT
    ws["N23"].number_format = PERCENT_FORMAT

    # 空行 24（间距）
    ws.merge_cells("L24:N24")

    # ── Notes / 备注 ──
    ws.merge_cells("A25:B25")
    ws["A25"] = "Notes/备注："
    ws["A25"].font = FONT_BODY_10
    ws["A25"].alignment = ALIGN_LEFT

    ws.merge_cells("A26:N26")
    ws["A26"] = "备注：\n" + "\n".join(NOTES)
    ws["A26"].font = FONT_BODY_10
    ws["A26"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 打印设置
    ws.print_area = "A1:N26"
    ws.page_setup.orientation = "landscape"

    return ws


# ─────────────────────────────────────────────
# Sheet 2: 开发工作量
# ─────────────────────────────────────────────

def build_workload_sheet(wb):
    ws = wb.create_sheet("开发工作量")

    # 列宽（与参考模板精确匹配）
    set_col_widths(ws, {
        "A": 2.38, "B": 6.12, "C": 16.12, "D": 52.88,
        "E": 11.0, "F": 9.62, "G": 12.88, "H": 38.25,
    })

    # 行高
    set_row_heights(ws, {1: 15.0, 2: 31.5, 3: 21.95})

    # 标题
    ws.merge_cells("B2:H2")
    ws["B2"] = "工作量评估表"
    ws["B2"].font = FONT_WL_TITLE
    ws["B2"].alignment = ALIGN_CENTER

    # 表头
    headers = ["ID", "用户场景", "任务名称", "PM/BA/SA", "PG/PT", "费用小计", "备注"]
    for col_idx, val in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.font = FONT_WL_BOLD
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # ── 按用户场景分组 ──
    from itertools import groupby
    groups = []
    for key, grp in groupby(WORKLOAD_ITEMS, key=lambda x: x["scenario"]):
        groups.append((key, list(grp)))

    data_start = 4
    unit_price = CONFIG["unit_price"]
    current_row = data_start
    item_id = 1

    for scenario, items in groups:
        group_start = current_row
        group_end = current_row + len(items) - 1

        # 写每一行数据
        for i, item in enumerate(items):
            row = current_row + i
            ws.cell(row=row, column=4, value=item["task_name"]).font = FONT_WL
            ws.cell(row=row, column=5, value=item["pm_ba_sa_days"] or None).font = FONT_WL
            ws.cell(row=row, column=6, value=item["pg_pt_days"] or None).font = FONT_WL
            ws.cell(row=row, column=8, value=item.get("remark", "")).font = FONT_WL

            # 边框和对齐
            for c in range(2, 9):
                ws.cell(row=row, column=c).border = THIN_BORDER
                ws.cell(row=row, column=c).alignment = ALIGN_CENTER if c in (2, 5, 6, 7) else ALIGN_LEFT

        # 用户场景列：首行写值，多行则合并
        ws.cell(row=group_start, column=3, value=scenario).font = FONT_WL
        ws.cell(row=group_start, column=3).alignment = ALIGN_CENTER
        if len(items) > 1:
            ws.merge_cells(start_row=group_start, start_column=3, end_row=group_end, end_column=3)

        # ID列：首行写值，多行则合并
        ws.cell(row=group_start, column=2, value=item_id).font = FONT_WL
        ws.cell(row=group_start, column=2).alignment = ALIGN_CENTER
        if len(items) > 1:
            ws.merge_cells(start_row=group_start, start_column=2, end_row=group_end, end_column=2)

        # 费用小计列：按场景组汇总，多行则合并
        fee_formula = f"=SUM(E{group_start}:F{group_end})*{unit_price}"
        ws.cell(row=group_start, column=7, value=fee_formula).font = FONT_WL
        ws.cell(row=group_start, column=7).number_format = MONEY_FORMAT_WL
        ws.cell(row=group_start, column=7).alignment = ALIGN_CENTER
        if len(items) > 1:
            ws.merge_cells(start_row=group_start, start_column=7, end_row=group_end, end_column=7)

        current_row = group_end + 1
        item_id += 1

    # 小计行
    subtotal_row = current_row
    ws.merge_cells(f"B{subtotal_row}:D{subtotal_row}")
    ws.cell(row=subtotal_row, column=2, value="小计：").font = FONT_WL
    ws.cell(row=subtotal_row, column=2).alignment = ALIGN_RIGHT
    ws.cell(row=subtotal_row, column=5, value=f"=SUM(E{data_start}:E{subtotal_row - 1})").font = FONT_WL
    ws.cell(row=subtotal_row, column=6, value=f"=SUM(F{data_start}:F{subtotal_row - 1})").font = FONT_WL
    style_range(ws, f"B{subtotal_row}:H{subtotal_row}", fill=FILL_HEADER, border=THIN_BORDER, alignment=ALIGN_CENTER)

    # 合计行
    total_row = subtotal_row + 1
    ws.merge_cells(f"B{total_row}:D{total_row}")
    ws.merge_cells(f"E{total_row}:F{total_row}")
    ws.cell(row=total_row, column=2, value="合计：").font = FONT_WL
    ws.cell(row=total_row, column=2).alignment = ALIGN_RIGHT
    ws.cell(row=total_row, column=7, value=f"=SUM(G{data_start}:G{subtotal_row - 1})").font = FONT_WL
    ws.cell(row=total_row, column=7).number_format = MONEY_FORMAT_WL
    ws.cell(row=total_row, column=8, value="不含税价").font = FONT_WL_BOLD
    ws.cell(row=total_row, column=8).alignment = ALIGN_LEFT
    style_range(ws, f"B{total_row}:H{total_row}", fill=FILL_HEADER, border=THIN_BORDER, alignment=ALIGN_CENTER)

    # 冻结表头
    ws.freeze_panes = "A4"

    return ws


# ─────────────────────────────────────────────
# 主入口
# ─────────────────────────────────────────────

def main():
    wb = Workbook()
    build_quote_sheet(wb)
    build_workload_sheet(wb)

    output_file = get_output_file_name()
    wb.save(output_file)
    print(f"已生成经典报价文件: {output_file}")

    # 校验
    unit_price = CONFIG["unit_price"]
    total_pm = sum(i["pm_ba_sa_days"] for i in WORKLOAD_ITEMS)
    total_pg = sum(i["pg_pt_days"] for i in WORKLOAD_ITEMS)
    total_days = total_pm + total_pg
    total_fee = total_days * unit_price
    total_fee_tax = total_fee * (1 + CONFIG["tax_rate"])

    print(f"  PM/BA/SA: {total_pm} 人天")
    print(f"  PG/PT:    {total_pg} 人天")
    print(f"  总人天:   {total_days} 人天")
    print(f"  单价:     ¥{unit_price:,}/人天")
    print(f"  不含税:   ¥{total_fee:,.2f}")
    print(f"  含税(6%): ¥{total_fee_tax:,.2f}")


if __name__ == "__main__":
    main()
