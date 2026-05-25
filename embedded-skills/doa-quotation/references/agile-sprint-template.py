# -*- coding: utf-8 -*-
"""
敏捷任务拆分 Excel 生成脚本骨架

用途：
- 生成 7 个工作表的敏捷拆分模板
- 默认参考 Sprint总览 / 任务明细拆分 / 工时校验 / 燃尽图 / 风险追踪 / 用户故事追溯 / Scrum仪式
- 适配通用商务蓝风格

使用方式：
1. 优先按 parameter-standard.md 中的标准字段替换 CONFIG、SPRINTS、TASKS 等占位数据
2. 按实际项目增减 Sprint 与任务
3. 运行：python agile-sprint-template.py
"""

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLOR_DARK_BLUE = "1F4E79"
COLOR_MID_BLUE = "2E75B6"
COLOR_LIGHT_BLUE = "D9EAF7"
COLOR_LIGHT_GRAY = "F5F7FA"
COLOR_WHITE = "FFFFFF"
COLOR_GREEN = "548235"
COLOR_ORANGE = "ED7D31"
COLOR_RED = "C00000"
COLOR_YELLOW = "FFC000"
MONEY_FORMAT = '¥#,##0'

HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color=COLOR_WHITE)
BODY_FONT = Font(name="微软雅黑", size=9)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color=COLOR_DARK_BLUE)
SUBTITLE_FONT = Font(name="微软雅黑", size=11, bold=True, color=COLOR_MID_BLUE)
HEADER_FILL = PatternFill(fill_type="solid", start_color=COLOR_DARK_BLUE, end_color=COLOR_DARK_BLUE)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

SPRINT_COLORS = {
    "Sprint 0": "D9E2F3",
    "Sprint 1": "E2EFDA",
    "Sprint 2": "FFF2CC",
    "Sprint 3": "FCE4EC",
    "Sprint 4": "FBE5D6",
    "Sprint 5": "D5B9DA",
}

PRIORITY_COLORS = {
    "P0": COLOR_RED,
    "P1": COLOR_ORANGE,
    "P2": COLOR_YELLOW,
}

CONFIG = {
    "output_file": "示例项目_敏捷任务拆分.xlsx",
    "project_name": "示例项目",
    "customer_name": "待填写客户",
    "vendor_name": "您的公司名称",
    "vendor_name_en": "",
    "template_source": "敏捷任务拆分模板",
    "currency_symbol": "¥",
    "delivery_cycle": "12周（约3个月）",
    "quote_owner": "待填写",
    "customer_contact": "待填写",
    "customer_email": "待填写",
    "customer_phone": "待填写",
    "remark_notes": [],
    "title": "示例项目 — 敏捷开发Sprint总览",
    "methodology": "Scrum + 持续交付",
    "iteration_cycle": "每Sprint 2周，共6个Sprint",
    "duration_text": "12周（约3个月）",
    "methodology": "Scrum + 持续交付",
    "amount_ex_tax": 121000,
    "amount_inc_tax": 128260,
    "tax_rate": "6%",
    "pm_days_total": 8,
    "pg_days_total": 47,
    "total_days": 55,
}


def get_output_file_name():
    return CONFIG.get("output_file") or f"{CONFIG['project_name']}_敏捷任务拆分.xlsx"

SPRINTS = [
    {
        "sprint": "Sprint 0",
        "name": "需求分析与技术设计",
        "period": "第1-2周\n(5个工作日)",
        "goal": "完成需求分析、原型设计、技术方案设计、开发环境搭建",
        "pm_days": 2,
        "pg_days": 3,
        "fee": 11000,
        "milestone": "需求文档\n原型图\n技术方案",
        "acceptance": "PO评审通过\n技术方案确认",
    },
    {
        "sprint": "Sprint 1",
        "name": "核心功能开发",
        "period": "第3-4周\n(10个工作日)",
        "goal": "完成首批核心功能和接口对接",
        "pm_days": 0,
        "pg_days": 8,
        "fee": 17600,
        "milestone": "核心模块可用",
        "acceptance": "核心接口打通",
    },
]

TASKS = [
    {
        "task_id": "T-0.1",
        "sprint": "Sprint 0",
        "quote_id": "1",
        "category": "需求分析",
        "sub_item": "需求分析",
        "task_name": "需求梳理",
        "description": "梳理范围、边界、验收标准",
        "role": "PM/BA",
        "pm_days": 1,
        "pg_days": 0,
        "fee": 0,  # TODO: 根据用户提供的人天单价计算
        "priority": "P0",
        "dependency": "无",
        "story_id": "ALL",
        "acceptance": "需求文档通过评审",
        "status": "未开始",
    },
    {
        "task_id": "T-1.1",
        "sprint": "Sprint 1",
        "quote_id": "2",
        "category": "功能开发",
        "sub_item": "核心模块",
        "task_name": "核心功能开发",
        "description": "实现首批功能与基础联调",
        "role": "PG/PT",
        "pm_days": 0,
        "pg_days": 2,
        "fee": 4400,
        "priority": "P0",
        "dependency": "T-0.1",
        "story_id": "1.1",
        "acceptance": "功能可用且可演示",
        "status": "未开始",
    },
]

RISKS = [
    ["R-01", "外部接口权限审批延迟", "Sprint 1", "高", "中", "提前申请权限并准备Mock数据", "PM", "待跟进", ""],
]

USER_STORIES = [
    ["1.1", "用户可使用核心功能", "T-1.1", "Sprint 1", 0, 2, 4400],
]

CEREMONIES = [
    ["Sprint Planning", "每Sprint第1天", "2-4小时", "PO + SM + Dev Team", "确定Sprint目标", "Sprint Backlog"],
    ["Daily Standup", "每个工作日", "15分钟", "SM + Dev Team", "同步进展与障碍", "障碍列表"],
]

DOD_ITEMS = [
    "代码已完成评审",
    "关键功能测试通过",
    "验收标准满足",
    "无P0/P1缺陷",
]


def set_col_widths(ws, widths):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def apply_header(ws, row, columns):
    for col in range(1, columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def apply_body(ws, start_row, end_row, columns, center_columns=None):
    center_columns = center_columns or []
    for row in range(start_row, end_row + 1):
        fill = PatternFill(fill_type="solid", start_color=COLOR_LIGHT_GRAY if row % 2 == 0 else COLOR_WHITE, end_color=COLOR_LIGHT_GRAY if row % 2 == 0 else COLOR_WHITE)
        for col in range(1, columns + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill.fill_type is None:
                cell.fill = fill
            cell.font = BODY_FONT
            cell.alignment = ALIGN_CENTER if col in center_columns else ALIGN_LEFT
            cell.border = THIN_BORDER


def build_sprint_overview(wb):
    ws = wb.active
    ws.title = "Sprint总览"
    ws.merge_cells("A1:I1")
    ws["A1"] = CONFIG["title"]
    ws["A1"].font = TITLE_FONT

    info_rows = [
        ["项目名称", CONFIG["project_name"], "", "报价总额(不含税)", CONFIG["amount_ex_tax"]],
        ["开发商", CONFIG["vendor_name"], "", "报价总额(含税)", CONFIG["amount_inc_tax"]],
        ["客户", CONFIG["customer_name"], "", "税率", CONFIG["tax_rate"]],
        ["迭代周期", CONFIG["iteration_cycle"], "", "PM/BA人天", CONFIG["pm_days_total"]],
        ["总周期", CONFIG["duration_text"], "", "PG/PT人天", CONFIG["pg_days_total"]],
        ["方法论", CONFIG["methodology"], "", "总人天", CONFIG["total_days"]],
    ]
    for idx, row in enumerate(info_rows, start=3):
        for col, value in [(1, row[0]), (2, row[1]), (4, row[3]), (5, row[4])]:
            ws.cell(row=idx, column=col, value=value)
            ws.cell(row=idx, column=col).font = BODY_FONT if col in (2, 5) else SUBTITLE_FONT

    start = 11
    headers = ["Sprint", "名称", "周期", "Sprint目标", "PM/BA\n人天", "PG/PT\n人天", "费用小计\n(不含税)", "交付里程碑", "验收标准"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=start, column=idx, value=header)
    apply_header(ws, start, len(headers))

    current = start + 1
    for sprint in SPRINTS:
        values = [
            sprint["sprint"], sprint["name"], sprint["period"], sprint["goal"], sprint["pm_days"], sprint["pg_days"], sprint["fee"], sprint["milestone"], sprint["acceptance"],
        ]
        for idx, value in enumerate(values, start=1):
            ws.cell(row=current, column=idx, value=value)
            if idx == 7:
                ws.cell(row=current, column=idx).number_format = MONEY_FORMAT
        sprint_fill = SPRINT_COLORS.get(sprint["sprint"])
        if sprint_fill:
            fill = PatternFill(fill_type="solid", start_color=sprint_fill, end_color=sprint_fill)
            for col in range(1, len(headers) + 1):
                ws.cell(row=current, column=col).fill = fill
        current += 1

    apply_body(ws, start + 1, current - 1, len(headers), center_columns=[1, 3, 5, 6, 7])
    set_col_widths(ws, {"A": 12, "B": 22, "C": 16, "D": 40, "E": 10, "F": 10, "G": 14, "H": 28, "I": 32})
    ws.freeze_panes = "A12"


def build_task_sheet(wb):
    ws = wb.create_sheet("任务明细拆分")
    ws.merge_cells("A1:P1")
    ws["A1"] = "敏捷任务明细拆分表"
    ws["A1"].font = TITLE_FONT

    headers = ["任务ID", "Sprint", "报价ID", "分类", "子项", "任务名称", "任务描述", "负责角色", "PM/BA\n人天", "PG/PT\n人天", "费用\n(不含税)", "优先级", "前置依赖", "用户故事ID", "验收标准", "状态"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=idx, value=header)
    apply_header(ws, 3, len(headers))

    row = 4
    for task in TASKS:
        values = [task["task_id"], task["sprint"], task["quote_id"], task["category"], task["sub_item"], task["task_name"], task["description"], task["role"], task["pm_days"], task["pg_days"], task["fee"], task["priority"], task["dependency"], task["story_id"], task["acceptance"], task["status"]]
        for idx, value in enumerate(values, start=1):
            ws.cell(row=row, column=idx, value=value)
            if idx == 11:
                ws.cell(row=row, column=idx).number_format = MONEY_FORMAT
        sprint_fill = SPRINT_COLORS.get(task["sprint"])
        if sprint_fill:
            fill = PatternFill(fill_type="solid", start_color=sprint_fill, end_color=sprint_fill)
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill
        priority_color = PRIORITY_COLORS.get(task["priority"])
        if priority_color:
            ws.cell(row=row, column=12).font = Font(name="微软雅黑", size=9, bold=True, color=priority_color)
        row += 1

    apply_body(ws, 4, row - 1, len(headers), center_columns=[1, 2, 3, 8, 9, 10, 11, 12, 14, 16])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:P{row - 1}"
    set_col_widths(ws, {"A": 9, "B": 11, "C": 8, "D": 10, "E": 14, "F": 22, "G": 48, "H": 10, "I": 9, "J": 9, "K": 12, "L": 8, "M": 14, "N": 11, "O": 36, "P": 9})


def build_effort_sheet(wb):
    ws = wb.create_sheet("工时校验")
    ws.merge_cells("A1:H1")
    ws["A1"] = "人天工时校验表"
    ws["A1"].font = TITLE_FONT

    headers = ["Sprint", "PM/BA\n人天", "PG/PT\n人天", "人天合计", "费用\n(不含税)", "累计费用", "对应报价ID", "校验说明"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=idx, value=header)
    apply_header(ws, 3, len(headers))

    cumulative_fee = 0
    row = 4
    for sprint in SPRINTS:
        total_days = sprint["pm_days"] + sprint["pg_days"]
        cumulative_fee += sprint["fee"]
        values = [sprint["sprint"], sprint["pm_days"], sprint["pg_days"], total_days, sprint["fee"], cumulative_fee, "待映射", "按Sprint汇总"]
        for idx, value in enumerate(values, start=1):
            ws.cell(row=row, column=idx, value=value)
            if idx in (5, 6):
                ws.cell(row=row, column=idx).number_format = MONEY_FORMAT
        row += 1

    apply_body(ws, 4, row - 1, len(headers), center_columns=[1, 2, 3, 4, 5, 6, 7])
    set_col_widths(ws, {"A": 12, "B": 10, "C": 10, "D": 10, "E": 14, "F": 14, "G": 18, "H": 36})


def build_burndown_sheet(wb):
    ws = wb.create_sheet("燃尽图")
    ws.merge_cells("A1:I1")
    ws["A1"] = "Sprint燃尽图数据"
    ws["A1"].font = TITLE_FONT

    headers = ["时间节点", "计划剩余\n人天", "Sprint交付\n人天", "累计交付\n人天", "完成比例", "PM/BA\n累计", "PG/PT\n累计", "累计费用\n(不含税)", "理想线"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=idx, value=header)
    apply_header(ws, 3, len(headers))

    total_days = CONFIG["total_days"]
    cumulative_days = 0
    cumulative_pm = 0
    cumulative_pg = 0
    cumulative_fee = 0
    row = 4
    steps = ["项目启动"] + [f"{s['sprint']} 结束" for s in SPRINTS]
    ideal_drop = total_days / max(len(SPRINTS), 1)
    for index, step in enumerate(steps):
        if index == 0:
            values = [step, total_days, 0, 0, "0%", 0, 0, 0, total_days]
        else:
            sprint = SPRINTS[index - 1]
            delivered = sprint["pm_days"] + sprint["pg_days"]
            cumulative_days += delivered
            cumulative_pm += sprint["pm_days"]
            cumulative_pg += sprint["pg_days"]
            cumulative_fee += sprint["fee"]
            remaining = max(total_days - cumulative_days, 0)
            completion = f"{round(cumulative_days / total_days * 100)}%" if total_days else "0%"
            ideal = round(max(total_days - ideal_drop * index, 0), 1)
            values = [step, remaining, delivered, cumulative_days, completion, cumulative_pm, cumulative_pg, cumulative_fee, ideal]
        for idx, value in enumerate(values, start=1):
            ws.cell(row=row, column=idx, value=value)
            if idx == 8:
                ws.cell(row=row, column=idx).number_format = MONEY_FORMAT
        row += 1

    apply_body(ws, 4, row - 1, len(headers), center_columns=list(range(1, 10)))
    chart = LineChart()
    chart.title = "Sprint燃尽图"
    chart.y_axis.title = "人天"
    chart.x_axis.title = "时间节点"
    data = Reference(ws, min_col=2, max_col=2, min_row=3, max_row=row - 1)
    ideal = Reference(ws, min_col=9, max_col=9, min_row=3, max_row=row - 1)
    categories = Reference(ws, min_col=1, min_row=4, max_row=row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.add_data(ideal, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, f"A{row + 2}")
    set_col_widths(ws, {"A": 16, "B": 12, "C": 12, "D": 12, "E": 10, "F": 10, "G": 10, "H": 14, "I": 10})


def build_risk_sheet(wb):
    ws = wb.create_sheet("风险追踪")
    ws.merge_cells("A1:I1")
    ws["A1"] = "风险与依赖追踪表"
    ws["A1"].font = TITLE_FONT
    headers = ["风险ID", "风险描述", "影响Sprint", "影响程度", "发生概率", "应对措施", "责任人", "状态", "更新日期"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=idx, value=header)
    apply_header(ws, 3, len(headers))
    row = 4
    for risk in RISKS:
        for idx, value in enumerate(risk, start=1):
            ws.cell(row=row, column=idx, value=value)
        row += 1
    apply_body(ws, 4, row - 1, len(headers), center_columns=[1, 3, 4, 5, 7, 8, 9])
    ws.auto_filter.ref = f"A3:I{row - 1}"
    set_col_widths(ws, {"A": 8, "B": 36, "C": 13, "D": 10, "E": 10, "F": 48, "G": 10, "H": 10, "I": 12})


def build_story_sheet(wb):
    ws = wb.create_sheet("用户故事追溯")
    ws.merge_cells("A1:G1")
    ws["A1"] = "用户故事 -> 任务追溯矩阵"
    ws["A1"].font = TITLE_FONT
    headers = ["用户故事ID", "用户故事描述", "关联任务ID", "关联Sprint", "PM/BA\n人天", "PG/PT\n人天", "费用\n(不含税)"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=idx, value=header)
    apply_header(ws, 3, len(headers))
    row = 4
    for story in USER_STORIES:
        for idx, value in enumerate(story, start=1):
            ws.cell(row=row, column=idx, value=value)
            if idx == 7:
                ws.cell(row=row, column=idx).number_format = MONEY_FORMAT
        row += 1
    apply_body(ws, 4, row - 1, len(headers), center_columns=[1, 4, 5, 6, 7])
    set_col_widths(ws, {"A": 12, "B": 36, "C": 28, "D": 18, "E": 10, "F": 10, "G": 14})


def build_scrum_sheet(wb):
    ws = wb.create_sheet("Scrum仪式")
    ws.merge_cells("A1:F1")
    ws["A1"] = "Scrum仪式日历"
    ws["A1"].font = TITLE_FONT
    headers = ["仪式名称", "频率", "时长", "参与人", "目的", "输出物"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=idx, value=header)
    apply_header(ws, 3, len(headers))
    row = 4
    for ceremony in CEREMONIES:
        for idx, value in enumerate(ceremony, start=1):
            ws.cell(row=row, column=idx, value=value)
        row += 1
    apply_body(ws, 4, row - 1, len(headers), center_columns=[2, 3])

    dod_row = row + 2
    ws.merge_cells(start_row=dod_row, start_column=1, end_row=dod_row, end_column=6)
    ws.cell(row=dod_row, column=1, value="Definition of Done")
    ws.cell(row=dod_row, column=1).font = SUBTITLE_FONT
    for offset, item in enumerate(DOD_ITEMS, start=1):
        ws.cell(row=dod_row + offset, column=1, value=f"- {item}")
    set_col_widths(ws, {"A": 22, "B": 18, "C": 10, "D": 22, "E": 42, "F": 20})


def main():
    workbook = Workbook()
    build_sprint_overview(workbook)
    build_task_sheet(workbook)
    build_effort_sheet(workbook)
    build_burndown_sheet(workbook)
    build_risk_sheet(workbook)
    build_story_sheet(workbook)
    build_scrum_sheet(workbook)
    output_file = get_output_file_name()
    workbook.save(output_file)
    print(f"已生成敏捷任务拆分文件: {output_file}")


if __name__ == "__main__":
    main()
